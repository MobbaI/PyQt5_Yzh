import sys
from PyQt5.QtWidgets import QApplication, QVBoxLayout,QHBoxLayout, QPushButton, QLineEdit, QWidget, QLabel, \
    QCheckBox, QFileDialog, QMessageBox
from PyQt5.QtCore import QThread, pyqtSignal
import openpyxl
import os


class DeleteNot(QWidget):
    def __init__(self):
        super(DeleteNot, self).__init__()
        self.fname = ''
        self.initUI()

    def initUI(self):
        self.setWindowTitle('删除未测试设备')
        self.setFixedSize(300, 110)

        hwidget_1 = QWidget()
        hwidget_2 = QWidget()

        self.lineedit = QLineEdit()
        self.lineedit.setReadOnly(True)
        self.button_1 = QPushButton('...')

        self.button_1.clicked.connect(self.on_click_button_1)

        hlayout_1 = QHBoxLayout(hwidget_1)
        hlayout_1.addWidget(self.lineedit)
        hlayout_1.addWidget(self.button_1)
        # hlayout_1.setStretch(0, 10)
        # hlayout_1.setStretch(1, 1)

        self.android = QCheckBox('安卓')
        self.ios = QCheckBox('iOS')
        self.label = QLabel()
        self.button_2 = QPushButton('运行')

        self.button_2.clicked.connect(self.delete_button_2)

        hlayout_2 = QHBoxLayout(hwidget_2)
        hlayout_2.addWidget(self.android)
        hlayout_2.addWidget(self.ios)
        hlayout_2.addWidget(self.label)
        hlayout_2.addWidget(self.button_2)

        vlayout = QVBoxLayout()
        vlayout.addWidget(hwidget_1)
        vlayout.addWidget(hwidget_2)

        self.setLayout(vlayout)

    def on_click_button_1(self):
        self.fname, _ = QFileDialog.getOpenFileName(self, '选择Excel文件', '.', '*.xlsx')
        self.lineedit.setText(os.path.basename(self.fname))

    def delete_button_2(self):
        if not(self.fname) or (self.android.checkState() == 0 and self.ios.checkState() == 0):
            QMessageBox.information(self, '提示', '请选择')
        else:
            self.button_1.setEnabled(False)
            self.button_2.setEnabled(False)
            self.deleterow = DeleteRow(self.fname, 3)
            self.deleterow.sinOut.connect(self.abc)
            self.deleterow.start()

    def abc(self, abc):
        if abc == 1:
            self.button_1.setEnabled(True)
            self.button_2.setEnabled(True)
            QMessageBox.information(self, '提示', '已完成')


class DeleteRow(QThread):
    sinOut = pyqtSignal(int)

    def __init__(self, fname, state):
        super(DeleteRow, self).__init__()
        self.fname = fname
        self.state = state

    def run(self):
        wb = openpyxl.load_workbook(self.fname)
        try:
            ws = wb['测试记录']
            if self.state == 1:
                self.android_delete(ws)
            if self.state == 2:
                self.ios_delete(ws)
            if self.state == 3:
                self.android_delete(ws)
                self.ios_delete(ws)
            fnamenew = self.fname.rstrip('.xlsx')
            wb.save('{}_new.xlsx'.format(fnamenew))
        except Exception:
            pass
        abc = 1
        self.sinOut.emit(abc)
        # QMessageBox.information(self, '提示', '已完成')

    def android_delete(self, ws):
        for cell in ws['B']:
            if cell.value == '安卓资产编号':
                android_begin_row = cell.row
            if cell.value == 'android_end_row':
                android_end_row = cell.row
        # 高风险项在第一行，不能删除
        android_o = [android_begin_row + 1, ]
        for i in range(android_begin_row + 1, android_end_row):
            if ws['O{}'.format(i)].value:
                android_o.append(ws['O{}'.format(i)].row)
        android_o.append(android_end_row)
        android_o_not = []
        for j in range(1, len(android_o)):
            android_o_not.append(android_o[j] - android_o[j - 1] - 1)
        begin = android_begin_row + 2
        for k in android_o_not:
            if k != 0:
                ws.delete_rows(begin, k)
            begin += 1

    def ios_delete(self, ws):
        for cell in ws['B']:
            if cell.value == 'iOS资产编号':
                ios_begin_row = cell.row
            if cell.value == 'ios_end_row':
                ios_end_row = cell.row
        # 高风险项在第一行，不能删除
        ios_o = [ios_begin_row + 1, ]
        for i in range(ios_begin_row + 1, ios_end_row):
            if ws['O{}'.format(i)].value:
                ios_o.append(ws['O{}'.format(i)].row)
        ios_o.append(ios_end_row)
        android_o_not = []
        for j in range(1, len(ios_o)):
            android_o_not.append(ios_o[j] - ios_o[j - 1] - 1)
        begin = ios_begin_row + 2
        for k in android_o_not:
            if k != 0:
                ws.delete_rows(begin, k)
            begin += 1


if __name__ == '__main__':
    app = QApplication(sys.argv)
    deletenot = DeleteNot()
    deletenot.show()
    sys.exit(app.exec_())