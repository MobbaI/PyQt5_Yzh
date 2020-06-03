import sys
from PyQt5.QtWidgets import QApplication, QVBoxLayout,QHBoxLayout, QPushButton, QLineEdit, QWidget, \
    QCheckBox, QFileDialog, QMessageBox
from PyQt5.QtCore import QThread, pyqtSignal
import openpyxl
import os
import re
import copy


num = 0


class ModifySymbolMain(QWidget):
    def __init__(self):
        super(ModifySymbolMain, self).__init__()
        self.fname = ''
        self.initUI()

    def initUI(self):
        self.setWindowTitle('修改标识符')
        self.setFixedSize(350, 120)

        hwidget_1 = QWidget()
        hwidget_2 = QWidget()

        self.lineedit = QLineEdit()
        self.lineedit.setReadOnly(True)
        self.button_1 = QPushButton('...')

        self.button_1.clicked.connect(self.on_click_button_1)

        hlayout_1 = QHBoxLayout(hwidget_1)
        hlayout_1.addWidget(self.lineedit)
        hlayout_1.addWidget(self.button_1)
        # hlayout_1.setStretch(0, 10)
        # hlayout_1.setStretch(1, 1)

        self.checkbox1 = QCheckBox('运行')
        self.checkbox2 = QCheckBox('UI')
        self.checkbox3 = QCheckBox('模型')
        self.checkbox4 = QCheckBox('场景')
        self.checkbox1.setCheckState(2)
        self.checkbox2.setCheckState(2)
        self.checkbox3.setCheckState(2)
        self.checkbox4.setCheckState(2)
        self.button_2 = QPushButton('运行')

        self.button_2.clicked.connect(self.on_click_button_2)

        hlayout_2 = QHBoxLayout(hwidget_2)
        hlayout_2.addWidget(self.checkbox1)
        hlayout_2.addWidget(self.checkbox2)
        hlayout_2.addWidget(self.checkbox3)
        hlayout_2.addWidget(self.checkbox4)
        hlayout_2.addWidget(self.button_2)

        vlayout = QVBoxLayout()
        vlayout.addWidget(hwidget_1)
        vlayout.addWidget(hwidget_2)

        self.setLayout(vlayout)

    def on_click_button_1(self):
        self.fname, _ = QFileDialog.getOpenFileName(self, '选择Excel文件', '.', '*.xlsx')
        self.lineedit.setText(os.path.basename(self.fname))

    def on_click_button_2(self):
        if not(self.fname) or (self.checkbox1.checkState() == 0 and self.checkbox2.checkState() == 0 and
                               self.checkbox3.checkState() == 0 and self.checkbox4.checkState() == 0):
            QMessageBox.information(self, '提示', '请选择')
        else:
            self.button_1.setEnabled(False)
            self.button_2.setEnabled(False)

            self.wb = openpyxl.load_workbook(self.fname)

            ws = self.wb['测试记录']

            self.modify1 = ModifySymbol(ws, 'U', '运行', self.checkbox1.checkState())
            self.modify2 = ModifySymbol(ws, 'W', 'UI', self.checkbox2.checkState())
            self.modify3 = ModifySymbol(ws, 'X', '模型问题', self.checkbox3.checkState())
            self.modify4 = ModifySymbol(ws, 'Y', '场景问题', self.checkbox4.checkState())
            self.finish = Finish()
            self.modify1.sinOut.connect(self.abc)
            self.modify2.sinOut.connect(self.abc2)
            self.modify2.sinOut.connect(self.abc3)
            self.modify2.sinOut.connect(self.abc4)
            self.finish.sinOut.connect(self.abc5)
            self.modify1.start()
            self.modify2.start()
            self.modify3.start()
            self.modify4.start()
            self.finish.start()

    def abc(self, abc):
        global num
        if abc == 'U':
            num += 1

    def abc2(self, abc):
        global num
        if abc == 'W':
            num += 1

    def abc3(self, abc):
        global num
        if abc == 'W':
            num += 1

    def abc4(self, abc):
        global num
        if abc == 'W':
            num += 1

    def abc5(self, abc):
        if abc == 1:
            self.button_1.setEnabled(True)
            self.button_2.setEnabled(True)
            fnamenew = self.fname.rstrip('.xlsx')
            self.wb.save('{}_new.xlsx'.format(fnamenew))
            QMessageBox.information(self, '提示', '已完成')


class ModifySymbol(QThread):
    sinOut = pyqtSignal(str)

    def __init__(self, ws, column, name, checkbox_state):
        super(ModifySymbol, self).__init__()
        self.ws = ws
        self.column = column
        self.name = name
        self.checkbox_state = checkbox_state

    def run(self):
        # wb = openpyxl.load_workbook(self.fname)
        if self.checkbox_state == 2:
            try:
                # ws = wb['测试记录']
                font = self.ws['Q12'].font
                alignment = self.ws['Q12'].alignment
                # fill = ws['Q17'].fill
                for i in range(17, 141):
                    cell = self.ws['{}{}'.format(self.column, i)]
                    if cell.value == 'U':
                        bug_name = self.ws['AQ{}'.format(i)].value
                        if bug_name:
                            pattern = '.*?.{}\(P.*?'.format(self.name)
                            result = re.match(pattern, bug_name)
                            if not result:
                                cell.value = '√'
                                cell.font = copy.copy(font)
                                cell.alignment = copy.copy(alignment)
                        else:
                            cell.value = '√'
                            cell.font = copy.copy(font)
                            cell.alignment = copy.copy(alignment)
                            # cell.fill = copy.copy(fill)
                # fnamenew = self.fname.rstrip('.xlsx')
                # wb.save('{}_new.xlsx'.format(fnamenew))
            except Exception as e:
                print(e)
        self.sinOut.emit(self.column)


class Finish(QThread):
    sinOut = pyqtSignal(int)

    def __init__(self):
        super(Finish, self).__init__()

    def run(self):
        global num
        while True:
            if num == 4:
                self.sinOut.emit(1)
                num = 0
                break
            self.sleep(1)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    modifysymbolmain = ModifySymbolMain()
    modifysymbolmain.show()
    sys.exit(app.exec_())