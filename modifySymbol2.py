import sys
from PyQt5.QtWidgets import QApplication, QVBoxLayout,QHBoxLayout, QPushButton, QLineEdit, QWidget, \
    QComboBox, QFileDialog, QMessageBox, QListView
from PyQt5.QtCore import QThread, pyqtSignal
import openpyxl
import os
import re
import copy


num = 0


class ModifySymbolMain(QWidget):
    def __init__(self):
        super(ModifySymbolMain, self).__init__()
        self.fname = ''
        self.initUI()

    def initUI(self):
        self.setWindowTitle('修改标识符')
        self.setFixedSize(350, 120)

        hwidget_1 = QWidget()
        hwidget_2 = QWidget()

        self.lineedit = QLineEdit()
        self.lineedit.setReadOnly(True)
        self.button_1 = QPushButton('...')

        self.button_1.clicked.connect(self.on_click_button_1)

        hlayout_1 = QHBoxLayout(hwidget_1)
        hlayout_1.addWidget(self.lineedit)
        hlayout_1.addWidget(self.button_1)
        # hlayout_1.setStretch(0, 10)
        # hlayout_1.setStretch(1, 1)

        self.combobox = QComboBox()

        self.combobox.setStyleSheet("QAbstractItemView::item {height: 23px;}")
        self.combobox.setView(QListView())

        self.combobox.addItems(['运行', '屏幕适配', 'UI', '模型问题', '场景问题', 'SDK', '输入法', '进程返回', '音频'])
        self.button_2 = QPushButton('运行')

        self.button_2.clicked.connect(self.on_click_button_2)

        hlayout_2 = QHBoxLayout(hwidget_2)
        hlayout_2.addWidget(self.combobox)
        hlayout_2.addWidget(self.button_2)
        hlayout_2.setStretch(0, 2)
        hlayout_2.setStretch(1, 1)

        vlayout = QVBoxLayout()
        vlayout.addWidget(hwidget_1)
        vlayout.addWidget(hwidget_2)

        self.setLayout(vlayout)

    def on_click_button_1(self):
        self.fname, _ = QFileDialog.getOpenFileName(self, '选择Excel文件', '.', '*.xlsx')
        self.lineedit.setText(os.path.basename(self.fname))

    def on_click_button_2(self):
        if not self.fname:
            QMessageBox.information(self, '提示', '请选择')
        else:
            self.button_1.setEnabled(False)
            self.button_2.setEnabled(False)
            self.combobox.setEnabled(False)
            self.modify = ModifySymbol(self.fname, self.combobox.currentText())
            self.modify.sinOut.connect(self.abc)
            self.modify.start()

    def abc(self, abc):
        if abc == 1:
            self.button_1.setEnabled(True)
            self.button_2.setEnabled(True)
            self.combobox.setEnabled(True)
            QMessageBox.information(self, '提示', '已完成')
        if abc == 2:
            QMessageBox.critical(self, '错误', '运行错误（可能表格未关闭）')


class ModifySymbol(QThread):
    sinOut = pyqtSignal(int)

    def __init__(self, fname, name):
        super(ModifySymbol, self).__init__()
        self.fname = fname
        self.name = name

    def run(self):
        wb = openpyxl.load_workbook(self.fname)
        try:
            ws = wb['测试记录']
            font = ws['Q12'].font
            alignment = ws['Q12'].alignment
            # fill = ws['Q17'].fill
            columns = {'运行': 'U', '屏幕适配': 'V', 'UI': 'W', '模型问题': 'X', '场景问题': 'Y', 'SDK': 'Z',
                       '输入法': 'AA', '进程返回': 'AC', '音频': 'AE'}
            for j in ws['B']:
                if j.value == 'ios_end_row':
                    end = j.row
                    break
            for i in range(17, end):
                cell = ws['{}{}'.format(columns[self.name], i)]
                if cell.value == 'U':
                    bug_name = ws['AQ{}'.format(i)].value
                    if bug_name:
                        pattern = '.*?.{}\(P.*?'.format(self.name)
                        result = re.match(pattern, bug_name)
                        if not result:
                            cell.value = '√'
                            cell.font = copy.copy(font)
                            cell.alignment = copy.copy(alignment)
                    else:
                        cell.value = '√'
                        cell.font = copy.copy(font)
                        cell.alignment = copy.copy(alignment)
                        # cell.fill = copy.copy(fill)
            wb.save(self.fname)
            self.sinOut.emit(1)
        except Exception as e:
            print(e)
            self.sinOut.emit(2)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    modifysymbolmain = ModifySymbolMain()
    modifysymbolmain.show()
    sys.exit(app.exec_())